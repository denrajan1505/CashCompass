import csv
import io
from datetime import datetime
from sqlalchemy.orm import Session

from app.models.expense import Expense
from app.models.user import User


def _get_all_expenses(db: Session, user: User) -> list:
    return (
        db.query(Expense)
        .filter(Expense.user_id == user.id)
        .order_by(Expense.expense_date.desc())
        .all()
    )


def export_csv(db: Session, user: User) -> bytes:
    expenses = _get_all_expenses(db, user)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Merchant", "Category", "Amount", "Currency", "Notes", "Source"])
    for e in expenses:
        writer.writerow([
            e.expense_date, e.merchant or "", e.category,
            e.amount, e.currency, e.notes or "", e.source,
        ])
    return buf.getvalue().encode("utf-8-sig")  # utf-8-sig for Excel compatibility


def export_xlsx(db: Session, user: User) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    expenses = _get_all_expenses(db, user)
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = ["Date", "Merchant", "Category", "Amount", "Currency", "Notes", "Source"]
    header_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, e in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=str(e.expense_date))
        ws.cell(row=row, column=2, value=e.merchant or "")
        ws.cell(row=row, column=3, value=e.category)
        ws.cell(row=row, column=4, value=float(e.amount))
        ws.cell(row=row, column=5, value=e.currency)
        ws.cell(row=row, column=6, value=e.notes or "")
        ws.cell(row=row, column=7, value=e.source)
        if row % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F3F2FF", end_color="F3F2FF", fill_type="solid")

    col_widths = [12, 22, 18, 12, 10, 30, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(db: Session, user: User) -> bytes:
    from fpdf import FPDF

    expenses = _get_all_expenses(db, user)
    total = sum(float(e.amount) for e in expenses)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Header
    pdf.set_fill_color(108, 99, 255)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 14, "CashCompass - Expense Report", align="C", fill=True, new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  User: {user.full_name}  |  Total: {user.preferred_currency} {total:,.2f}",
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Table header
    col_w = [28, 42, 34, 26, 18, 22]
    headers = ["Date", "Merchant", "Category", "Amount", "Currency", "Source"]
    pdf.set_fill_color(108, 99, 255)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    for w, h in zip(col_w, headers):
        pdf.cell(w, 9, h, border=0, align="C", fill=True)
    pdf.ln()

    # Rows
    pdf.set_font("Helvetica", "", 8)
    for i, e in enumerate(expenses):
        pdf.set_fill_color(243, 242, 255) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(30, 30, 30)
        fill = True
        merchant = (e.merchant or "")[:22]
        pdf.cell(col_w[0], 8, str(e.expense_date), border=0, fill=fill)
        pdf.cell(col_w[1], 8, merchant, border=0, fill=fill)
        pdf.cell(col_w[2], 8, e.category[:20], border=0, fill=fill)
        pdf.cell(col_w[3], 8, f"{float(e.amount):,.2f}", border=0, align="R", fill=fill)
        pdf.cell(col_w[4], 8, e.currency, border=0, align="C", fill=fill)
        pdf.cell(col_w[5], 8, e.source, border=0, align="C", fill=fill)
        pdf.ln()

    # Footer total
    pdf.ln(4)
    pdf.set_fill_color(108, 99, 255)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 10, f"Total: {user.preferred_currency} {total:,.2f}  ({len(expenses)} transactions)",
             align="R", fill=True, new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())
